import requests
import json
import os
import fitz     # 对的，这是 PyMuPDF 的真实名字
from docx import Document
from openpyxl import load_workbook
import tkinter as tk
from tkinter import Listbox, Scrollbar, messagebox
from datetime import datetime
import threading

# ===================== 配置区域 =====================
API_KEY = "sk-8450ec174722497c9c3e58cf3dbf7f5e"        # 👈 替换成你的真实 API Key
DEFAULT_MODEL = "deepseek-v4-pro"
HISTORY_DIR = "history"         # 历史记录保存目录
# ===================================================
# 历史列表缓存（全局变量）
_history_cache = []          # 存储 [(mtime, path, display), ...]
_cache_ready = False         # 缓存是否已准备好
_cache_lock = threading.Lock()

def pre_scan_history():
    """后台预扫描历史文件，填充缓存"""
    global _history_cache, _cache_ready
    files = []
    for f in os.listdir(HISTORY_DIR):
        if f.endswith('.json'):
            full_path = os.path.join(HISTORY_DIR, f)
            try:
                mtime = os.path.getmtime(full_path)
                # 从文件名生成显示名称
                time_str = f.replace('chat_', '').replace('.json', '')
                display = f"对话 {time_str}"
            except:
                display = f
            files.append((mtime, full_path, display))
    files.sort(reverse=True)  # 最新在前
    with _cache_lock:
        _history_cache[:] = files
        _cache_ready = True

# 创建历史目录
os.makedirs(HISTORY_DIR, exist_ok=True)

url = "https://api.deepseek.com/chat/completions"
headers = {
    "Content-Type": "application/json",
    "Authorization": f"Bearer {API_KEY}"
}

# 全局配置
thinking_enabled = True
reasoning_effort = "medium"
temperature = 0.7
messages = []          # 当前对话消息列表

# ===================== 历史记录函数 =====================
def save_conversation():
    """保存当前对话到文件"""
    if not messages:
        print("没有对话内容，未保存")
        return
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = os.path.join(HISTORY_DIR, f"chat_{timestamp}.json")
    with open(filename, 'w', encoding='utf-8') as f:
        json.dump(messages, f, ensure_ascii=False, indent=2)
    print(f"✅ 对话已保存到 {filename}")

def load_conversation(filepath):
    """从文件加载对话历史"""
    global messages
    try:
        with open(filepath, 'r', encoding='utf-8') as f:
            loaded = json.load(f)
        if isinstance(loaded, list):
            messages = loaded
            print(f"✅ 已加载 {len(messages)} 条消息历史")
            return True
        else:
            print("❌ 文件格式错误")
            return False
    except Exception as e:
        print(f"❌ 加载失败: {e}")
        return False

def list_history_files():
    """返回历史文件列表（优先从缓存读取）"""
    with _cache_lock:
        if _cache_ready and _history_cache:
            return _history_cache.copy()
    # 如果缓存未就绪，则降级为实时扫描（一般不会发生）
    files = []
    for f in os.listdir(HISTORY_DIR):
        if f.endswith('.json'):
            full_path = os.path.join(HISTORY_DIR, f)
            mtime = os.path.getmtime(full_path)
            try:
                time_str = f.replace('chat_', '').replace('.json', '')
                display = f"对话 {time_str}"
            except:
                display = f
            files.append((mtime, full_path, display))
    files.sort(reverse=True)
    return files

def open_history_window():
    """打开Tkinter窗口选择历史对话"""
    files = list_history_files()
    if not files:
        messagebox.showinfo("提示", "没有找到任何历史对话记录")
        return

    # 创建窗口
    win = tk.Tk()
    win.title("选择要加载的历史对话")
    win.geometry("400x300")

    tk.Label(win, text="双击任意一条记录即可加载", pady=5).pack()

    frame = tk.Frame(win)
    frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)

    scrollbar = Scrollbar(frame)
    scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

    listbox = Listbox(frame, yscrollcommand=scrollbar.set, font=("微软雅黑", 10))
    listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
    scrollbar.config(command=listbox.yview)

    # 插入数据
    for _, path, display in files:
        listbox.insert(tk.END, display)
        # 将路径存储在listbox的item中（用字典映射）
        if not hasattr(listbox, 'paths'):
            listbox.paths = []
        listbox.paths.append(path)

    def on_double_click(event):
        selection = listbox.curselection()
        if not selection:
            return
        idx = selection[0]
        path = listbox.paths[idx]
        win.destroy()
        # 在主程序中加载
        load_conversation(path)
        print("现在可以继续聊天了。")

    listbox.bind('<Double-Button-1>', on_double_click)
    win.mainloop()

# ===================== 文件读取函数 =====================
def read_pdf(filepath):
    doc = fitz.open(filepath)
    text = ""
    for page in doc:
        text += page.get_text()
    doc.close()
    return text

def read_docx(filepath):
    doc = Document(filepath)
    text = "\n".join([para.text for para in doc.paragraphs])
    return text

def read_excel(filepath):
    wb = load_workbook(filepath, data_only=True)
    sheet = wb.active
    rows = []
    for row in sheet.iter_rows(values_only=True):
        row_text = ", ".join([str(cell) for cell in row if cell is not None])
        rows.append(row_text)
    wb.close()
    return "\n".join(rows)

def read_txt(filepath):
    encodings = ['utf-8', 'gbk', 'gb2312', 'latin-1']
    for enc in encodings:
        try:
            with open(filepath, 'r', encoding=enc) as f:
                return f.read()
        except UnicodeDecodeError:
            continue
    raise Exception("无法解码文件")

def load_file(filepath):
    if not os.path.exists(filepath):
        return None, f"文件不存在: {filepath}"
    ext = os.path.splitext(filepath)[1].lower()
    try:
        if ext == '.pdf':
            content = read_pdf(filepath)
        elif ext == '.docx':
            content = read_docx(filepath)
        elif ext in ['.xlsx', '.xls']:
            content = read_excel(filepath)
        elif ext == '.txt':
            content = read_txt(filepath)
        else:
            return None, f"不支持的文件类型: {ext}"
        if not content.strip():
            return None, "文件内容为空"
        max_chars = 50000
        if len(content) > max_chars:
            content = content[:max_chars] + "\n...(内容过长，已截断)"
        return content, None
    except Exception as e:
        return None, f"读取失败: {type(e).__name__} - {e}"

# ===================== 辅助函数 =====================
def print_help():
    print("\n📖 可用命令：")
    print("  /reset      - 清空当前对话记忆")
    print("  /mode       - 切换思考模式 (开/关)")
    print("  /depth [l/m/h] - 设置思考深度")
    print("  /temp [0-1] - 设置温度（仅关闭思考模式时有效）")
    print("  /file <路径> - 读取文件（.txt/.pdf/.docx/.xlsx）")
    print("  /save       - 手动保存当前对话")
    print("  /history    - 打开历史窗口，加载之前的对话")
    print("  /status     - 查看当前配置")
    print("  /help       - 显示本帮助")
    print("  exit/quit   - 退出并自动保存对话\n")

def show_status():
    print("\n🔧 当前配置：")
    print(f"  模型: {DEFAULT_MODEL}")
    print(f"  思考模式: {'开启' if thinking_enabled else '关闭'}")
    if thinking_enabled:
        print(f"  思考深度: {reasoning_effort}")
    else:
        print(f"  温度: {temperature}")
    print(f"  当前对话消息数: {len(messages)} 条\n")

def ask_deepseek(user_input):
    messages.append({"role": "user", "content": user_input})
    data = {
        "model": DEFAULT_MODEL,
        "messages": messages,
        "stream": False
    }
    if thinking_enabled:
        data["thinking"] = {"type": "enabled"}
        data["reasoning_effort"] = reasoning_effort
    else:
        data["thinking"] = {"type": "disabled"}
        data["temperature"] = temperature
    try:
        response = requests.post(url, headers=headers, json=data, timeout=90)
        if response.status_code == 200:
            result = response.json()
            message = result['choices'][0]['message']
            reasoning = message.get('reasoning_content', '')
            answer = message.get('content', '')
            messages.append({"role": "assistant", "content": answer})
            usage = result.get('usage', {})
            return reasoning, answer, None, usage
        else:
            err_msg = f"HTTP {response.status_code}: {response.text}"
            return None, None, err_msg, {}
    except Exception as e:
        return None, None, f"错误: {e}", {}

# ===================== 主程序 =====================
def main():
    print("=" * 60)
    print("🤖 DeepSeek 增强版 (自动保存历史 + 历史窗口)")
    print("💬 支持多轮对话、文件上传、思考模式")
    print("📖 输入 /help 查看所有命令")
    print("=" * 60)
    show_status()
    # 启动后台预扫描历史记录
    threading.Thread(target=pre_scan_history, daemon=True).start()

    # 如果存在上一次未正常保存的临时文件，可以尝试恢复（略）
    try:
        while True:
            user_input = input("\n你：").strip()
            if not user_input:
                continue

            # 退出命令
            if user_input.lower() in ["exit", "quit"]:
                save_conversation()
                print("👋 再见！")
                break

            elif user_input == "/help":
                print_help()
            elif user_input == "/reset":
                messages.clear()
                print("✅ 当前对话已清空")
            elif user_input == "/mode":
                thinking_enabled = not thinking_enabled
                print(f"✅ 思考模式已{'开启' if thinking_enabled else '关闭'}")
            elif user_input.startswith("/depth"):
                parts = user_input.split()
                if len(parts) == 2 and parts[1].lower() in ['low', 'medium', 'high']:
                    reasoning_effort = parts[1].lower()
                    print(f"✅ 思考深度: {reasoning_effort}")
                else:
                    print("❌ 用法: /depth low|medium|high")
            elif user_input.startswith("/temp"):
                parts = user_input.split()
                if len(parts) == 2:
                    try:
                        t = float(parts[1])
                        if 0 <= t <= 2:
                            temperature = t
                            print(f"✅ 温度: {temperature}")
                        else:
                            print("❌ 温度应在 0~2")
                    except:
                        print("❌ 请输入数字")
                else:
                    print("❌ 用法: /temp 0.7")
            elif user_input.startswith("/file"):
                parts = user_input.split(maxsplit=1)
                if len(parts) != 2:
                    print("❌ 用法: /file C:\\path\\to\\file.pdf")
                    continue
                filepath = parts[1].strip()
                content, err = load_file(filepath)
                if err:
                    print(f"❌ {err}")
                else:
                    file_prompt = f"【用户上传的文件内容】\n```\n{content}\n```\n请基于以上文件内容回答后续问题。"
                    messages.append({"role": "user", "content": file_prompt})
                    print(f"✅ 已读取文件（{len(content)} 字符），现在可以提问")
            elif user_input == "/save":
                save_conversation()
            elif user_input == "/history":
                open_history_window()
            elif user_input == "/status":
                show_status()
            else:
                print("🤔 思考中...")
                reasoning, answer, err, usage = ask_deepseek(user_input)
                if err:
                    print(f"❌ {err}")
                else:
                    if reasoning:
                        print("\n🧠 思考过程：")
                        print(reasoning)
                    print("\n✨ 回答：")
                    print(answer)
                    if usage:
                        print(f"\n📊 Token消耗：输入 {usage.get('prompt_tokens',0)}，输出 {usage.get('completion_tokens',0)}，合计 {usage.get('total_tokens',0)}")
    except KeyboardInterrupt:
        print("\n检测到强制退出，正在保存对话...")
        save_conversation()

if __name__ == "__main__":
    main()